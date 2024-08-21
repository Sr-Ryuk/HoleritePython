import webbrowser
from openpyxl import Workbook, workbook, load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.styles import Alignment

planilha = Workbook()
guiaAtiva = planilha.active

guiaAtiva["A1"] = "Nome"
guiaAtiva["B1"] = "Cargo"
guiaAtiva["C1"] = "Salário"
guiaAtiva["D1"] = "INSS"
guiaAtiva["E1"] = "FGTS"
guiaAtiva["F1"] = "IRRF"
guiaAtiva["G1"] = "VT"
guiaAtiva["H1"] = "VA"
guiaAtiva["I1"] = "Data"
guiaAtiva["J1"] = "Hora"

my_fill = PatternFill(start_color = "5cb800", end_color = "5cb800", fill_type = "solid")
my_font = Font(bold = True)
my_header = ["A1", "B1", "C1", "D1", "E1", "F1", "G1", "H1", "I1", "J1"]
for cell in my_header:
    guiaAtiva[cell].fill = my_fill
    guiaAtiva[cell].font = my_font
    guiaAtiva[cell].alignment = Alignment(horizontal = "center")

guiaAtiva.row_dimensions[1].height = 15
guiaAtiva.column_dimensions["A"].width = 25
guiaAtiva.column_dimensions["B"].width = 25
guiaAtiva.column_dimensions["C"].width = 15
guiaAtiva.column_dimensions["D"].width = 15
guiaAtiva.column_dimensions["E"].width = 15
guiaAtiva.column_dimensions["F"].width = 15
guiaAtiva.column_dimensions["G"].width = 15
guiaAtiva.column_dimensions["H"].width = 15
guiaAtiva.column_dimensions["I"].width = 15
guiaAtiva.column_dimensions["J"].width = 15


def EnviarPlanilha(nome,cargo,valor,inss,irrf,fgts,valeT,valeA,data):
    nomePlanilha = "Pasta1.xlsx"
    planilha = load_workbook(nomePlanilha)
    guiaAtiva = planilha.active

    linha = len(list(guiaAtiva.rows))
    print(linha + 1)
    guiaAtiva[f"A{linha + 1}"] = nome
    guiaAtiva[f"B{linha + 1}"] = cargo
    guiaAtiva[f"C{linha + 1}"] = valor
    guiaAtiva[f"D{linha + 1}"] = inss
    guiaAtiva[f"E{linha + 1}"] = irrf
    guiaAtiva[f"F{linha + 1}"] = fgts
    guiaAtiva[f"G{linha + 1}"] = valeT
    guiaAtiva[f"H{linha + 1}"] = valeA
    guiaAtiva[f"I{linha + 1}"] = data[0]
    guiaAtiva[f"J{linha + 1}"] = data[1]
    planilha.save(nomePlanilha)